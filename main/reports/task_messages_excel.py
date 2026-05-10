from openpyxl import Workbook
from datetime import datetime
from typing import List, NamedTuple, Any, Optional

from openpyxl.styles import Alignment

from main.api_client.schemas.task import Task


class MessageField(NamedTuple):
    name: str
    label: str
    getter: str


MESSAGE_FIELDS = [
    MessageField("id", "ID сообщения", "id"),
    MessageField("created_at", "Дата сообщения", "created_at"),
    MessageField("user.id", "ID автора", "user.id"),
    MessageField("user.name", "Имя автора", "user.name"),
    MessageField("user.email", "Email автора", "user.email"),
    MessageField("user.phone", "Телефон автора", "user.phone"),
]


def get_field_value(obj: Any, field_path: str) -> Any:
    parts = field_path.split(".")
    current = obj
    for part in parts:
        if current is None:
            return None
        if not hasattr(current, part):
            return None
        current = getattr(current, part)
    return current


def create_task_messages_excel(
        task: Task,
        include_message_columns: Optional[List[str]] = None,
) -> Workbook:

    wb = Workbook()
    ws = wb.active
    ws.title = "Сообщения задачи"

    if include_message_columns is None:
        selected_fields = []
    else:
        selected_fields = [f for f in MESSAGE_FIELDS if f.name in include_message_columns]

    headers = ["Текст сообщения"]
    headers.extend([f.label for f in selected_fields])
    ws.append(headers)

    for message in task.messages:
        if message.message != "":
            row = [message.message]
        else:
            continue
        for field in selected_fields:
            value = get_field_value(message, field.getter)
            row.append(value)
        ws.append(row)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = None

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 50
    for col_num in range(2, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = 40

    return wb