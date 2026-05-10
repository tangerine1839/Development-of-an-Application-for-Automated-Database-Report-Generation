from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import List, NamedTuple, Any, Optional

from main.api_client.schemas.inspection import Inspection


class InspectionItemField(NamedTuple):
    name: str
    label: str
    getter: str


INSPECTION_ITEM_FIELDS = [
    InspectionItemField("description", "Описание", "description"),
    InspectionItemField("photos", "Фотографии", "photos"),
    InspectionItemField("value", "Ответ", "value"),
    InspectionItemField("triggers", "Триггеры", "triggers"),
]

CONDITION_TRANSLATION = {
    "ANY": "Любой ответ",
    "EQ": "Равно",
    "NOT_EQ": "Не равно",
    "ONE_OF": "Один из",
    "NOT_ONE_OF": "Не один из",
    "LESS": "Меньше",
    "LESS_OR_EQ": "Меньше или равно",
    "GREATER": "Больше",
    "GREATER_OR_EQ": "Больше или равно",
    "BETWEEN": "Между",
    "NOT_BETWEEN": "Не между",
}

ACTION_TRANSLATION = {
    "SHOW_ITEM": "Показать пункт",
    "REQUIRE_PHOTOS": "Требовать фото",
    "REQUIRE_BARCODES": "Требовать штрихкоды",
    "REQUIRE_AUDIOS": "Требовать аудио",
    "REQUIRE_VIDEOS": "Требовать видео",
    "REQUIRE_GEO": "Требовать геопозицию",
    "REQUIRE_COMMENT": "Требовать комментарий",
    "SET_IS_VIOLATED": "Считать нарушением",
    "IGNORE_RATE": "Игнорировать вес",
}


def get_field_value(obj: Any, field_path: str) -> Any:
    parts = field_path.split(".")
    current = obj
    for part in parts:
        if not hasattr(current, part):
            return None
        current = getattr(current, part)
        if current is None:
            break
    return current


def get_photos_urls(item: Any) -> str:
    photos = get_field_value(item, "photos")
    if not photos:
        return ""
    urls = []
    for photo in photos:
        url = get_field_value(photo, "url")
        if url:
            urls.append(url)
    return ", ".join(urls)


def check_condition(condition: str, item_value: Any, values: List[str]) -> bool:
    if not condition:
        return False

    if condition == "ANY":
        return True

    if condition == "EQ":
        return str(item_value) == values[0] if values else False

    if condition == "NOT_EQ":
        return str(item_value) != values[0] if values else False

    if condition == "ONE_OF":
        return str(item_value) in values if values else False

    if condition == "NOT_ONE_OF":
        return str(item_value) not in values if values else False

    if condition == "LESS":
        try:
            return float(item_value) < float(values[0]) if values else False
        except (ValueError, TypeError):
            return False

    if condition == "LESS_OR_EQ":
        try:
            return float(item_value) <= float(values[0]) if values else False
        except (ValueError, TypeError):
            return False

    if condition == "GREATER":
        try:
            return float(item_value) > float(values[0]) if values else False
        except (ValueError, TypeError):
            return False

    if condition == "GREATER_OR_EQ":
        try:
            return float(item_value) >= float(values[0]) if values else False
        except (ValueError, TypeError):
            return False

    if condition == "BETWEEN":
        try:
            return len(values) >= 2 and float(values[0]) <= float(item_value) <= float(values[1])
        except (ValueError, TypeError):
            return False

    if condition == "NOT_BETWEEN":
        try:
            return len(values) >= 2 and (float(item_value) < float(values[0]) or float(item_value) > float(values[1]))
        except (ValueError, TypeError):
            return False

    return False


def get_trigger_actions(item: Any) -> str:
    triggers = get_field_value(item, "triggers")
    if not triggers:
        return ""

    item_value = get_field_value(item, "value") or ""

    all_actions = []
    for trigger in triggers:
        condition = get_field_value(trigger, "condition") or ""
        values = get_field_value(trigger, "values") or []

        if check_condition(condition, item_value, values):
            actions = get_field_value(trigger, "actions") or []
            for action in actions:
                action_type = get_field_value(action, "type") or ""
                action_ru = ACTION_TRANSLATION.get(action_type, action_type)

                evidence_count = get_field_value(action, "evidence_count")
                if evidence_count and action_type in ["REQUIRE_PHOTOS", "REQUIRE_BARCODES", "REQUIRE_AUDIOS",
                                                      "REQUIRE_VIDEOS", "REQUIRE_GEO"]:
                    action_ru = f"{action_ru} ({evidence_count} шт.)"

                item_id = get_field_value(action, "item_id")
                if item_id and action_type == "SHOW_ITEM":
                    action_ru = f"{action_ru} (ID: {item_id})"

                all_actions.append(action_ru)

    return ", ".join(all_actions)


def create_inspection_items_excel(
        inspection: Inspection,
        include_columns: Optional[List[str]] = None,
) -> Workbook:


    wb = Workbook()
    ws = wb.active
    ws.title = "Пункты проверки"

    if include_columns is None:
        selected_fields = []
    else:
        selected_fields = [f for f in INSPECTION_ITEM_FIELDS if f.name in include_columns]

    headers = ["Наименование проверки"]
    headers.extend([f.label for f in selected_fields])
    ws.append(headers)

    items = get_field_value(inspection, "items") or []

    for item in items:
        item_name = get_field_value(item, "name") or ""
        item_type = get_field_value(item, "type")

        row = [item_name]

        if item_type in ["Section", "Category"]:
            for _ in selected_fields:
                row.append("")
        else:
            for field in selected_fields:
                if field.name == "description":
                    value = get_field_value(item, "description") or ""
                    row.append(value)
                elif field.name == "photos":
                    value = get_photos_urls(item)
                    row.append(value)
                elif field.name == "value":
                    value = get_field_value(item, "value") or ""
                    row.append(value)
                elif field.name == "triggers":
                    value = get_trigger_actions(item)
                    row.append(value)
                else:
                    row.append("")

        ws.append(row)

        current_row = ws.max_row

        ws.row_dimensions[current_row].height = None

        name_cell = ws.cell(row=current_row, column=1)
        name_cell.alignment = Alignment(wrap_text=True, vertical='top')

        if item_type == "Section":
            name_cell.font = Font(bold=True, size=12)
            name_cell.value = item_name.upper()
            name_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif item_type == "Category":
            name_cell.font = Font(bold=True, size=11)
            name_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            name_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(wrap_text=True, vertical='top')


    ws.column_dimensions['A'].width = 50
    for col_num in range(2, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = 40

    return wb