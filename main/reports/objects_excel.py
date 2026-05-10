from openpyxl import Workbook
from typing import List, NamedTuple, Any, Optional

from openpyxl.styles import Alignment

from main.api_client.schemas.place import Place


class ObjectField(NamedTuple):
    name: str
    label: str
    getter: str


OBJECT_FIELDS = [
    ObjectField("id", "ID", "id"),
    ObjectField("name", "Название", "name"),
    ObjectField("address", "Адрес", "address"),
    ObjectField("geo.lat", "Широта", "geo.lat"),
    ObjectField("geo.lng", "Долгота", "geo.lng"),
    ObjectField("code", "Код", "code"),
    ObjectField("zone1.name", "Зона 1", "zone1.name"),
    ObjectField("zone2.name", "Зона 2", "zone2.name"),
    ObjectField("zone3.name", "Зона 3", "zone3.name"),
    ObjectField("groups", "Группы", "groups"),
    ObjectField("manager", "Менеджер", "manager"),
]


def get_field_value(place: Place, field_path: str) -> Any:
    parts = field_path.split(".")
    current = place
    for part in parts:
        if not hasattr(current, part):
            return None
        current = getattr(current, part)
        if current is None:
            break
    return current


def create_objects_excel(
    objects: List[Place],
    columns: Optional[List[str]] = None,
    heading: str = "Отчёт по объектам",
) -> Workbook:

    if columns is None:
        columns = [f.name for f in OBJECT_FIELDS]

    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты"

    headers = [f.label for f in OBJECT_FIELDS if f.name in columns]
    ws.append(headers)

    field_paths = [f.getter for f in OBJECT_FIELDS if f.name in columns]

    for place in objects:
        row = []
        for path in field_paths:
            value = get_field_value(place, path)
            if isinstance(value, list) and path == "groups":
                # склеиваем названия групп через запятую
                value = ", ".join([g.name for g in value]) if value else ""
            elif value is None:
                value = ""
            row.append(value)
        ws.append(row)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = None

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 50
    for col_num in range(2, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = 40

    return wb