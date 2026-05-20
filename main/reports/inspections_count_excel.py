from openpyxl import Workbook
from datetime import date
from typing import List, Dict, Any

from openpyxl.styles import Alignment, Font


class PeriodCountField:
    def __init__(self, name: str, label: str, getter: str):
        self.name = name
        self.label = label
        self.getter = getter


PERIOD_COUNT_FIELDS = [
    PeriodCountField("period", "Период", "period"),
    PeriodCountField("count", "Количество проверок", "count"),
]


def get_field_value(item: Dict[str, Any], field_path: str) -> Any:
    parts = field_path.split(".")
    current = item
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        else:
            if not hasattr(current, part):
                return None
            current = getattr(current, part)
        if current is None:
            break
    return current


def format_period(period: date, grouping: str) -> str:
    if grouping == 'day':
        return period.strftime('%d.%m.%Y')
    else:
        return period.strftime('%B %Y')


def create_inspections_count_excel(
        data: List[Dict[str, Any]],
        grouping: str,
        date_from: date,
        date_to: date,
        heading: str = "Отчёт по количеству проверок",
) -> Workbook:


    wb = Workbook()
    ws = wb.active
    ws.title = "Количество проверок"

    # Заголовок отчёта
    ws.append([heading])
    ws.append([f"Период: с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"])
    ws.append([f"Группировка: {'по дням' if grouping == 'day' else 'по месяцам'}"])
    ws.append([])  # Пустая строка

    headers = [f.label for f in PERIOD_COUNT_FIELDS]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    total_count = 0
    for item in data:
        period = get_field_value(item, "period")
        count = get_field_value(item, "count")

        period_str = format_period(period, grouping)
        total_count += count

        row = [period_str, count]
        ws.append(row)

        current_row = ws.max_row
        ws.row_dimensions[current_row].height = None

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            if col_num == 1:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    ws.append([])
    ws.append(["ИТОГО:", total_count])

    total_row = ws.max_row
    for col_num in range(1, 3):
        cell = ws.cell(row=total_row, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center' if col_num == 2 else 'left', vertical='center', wrap_text=True)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    return wb