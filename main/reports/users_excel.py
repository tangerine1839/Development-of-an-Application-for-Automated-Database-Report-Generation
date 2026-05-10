from openpyxl import Workbook
from openpyxl.styles import Alignment
from typing import List, Optional, NamedTuple

from main.api_client.responses.users_list import UserListResponse
from main.api_client.schemas.user import User


class UserField(NamedTuple):
    name: str
    label: str
    getter: str


USER_FIELDS = [
    UserField("id", "ID пользователя", "id"),
    UserField("email", "Email", "email"),
    UserField("name", "Имя", "name"),
    UserField("role", "Роль", "role"),
    UserField("permission_group", "Группа прав", "permissionGroup.name"),
    UserField("position", "Должность", "position"),
    UserField("phone", "Телефон", "phone"),
    UserField("places_groups", "Группы объектов проверки", "placesGroups"),
    UserField("user_groups", "Группы", "groups"),
]


def get_field_value(obj, field_path: str):
    parts = field_path.split(".")
    current = obj
    for part in parts:
        if current is None:
            return None
        if not hasattr(current, part):
            return None
        current = getattr(current, part)
    return current


def get_places_groups_names(user: User) -> str:
    places_groups = getattr(user, "placesGroups", [])
    if not places_groups:
        return ""
    names = [getattr(group, "name", "") for group in places_groups if getattr(group, "name", None)]
    return ", ".join(names)


def get_user_groups_names(user: User) -> str:
    groups = getattr(user, "groups", [])
    if not groups:
        return ""
    names = [getattr(group, "name", "") for group in groups if getattr(group, "name", None)]
    return ", ".join(names)


def create_users_excel(
        users: List[User],
        include_columns: Optional[List[str]] = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    if include_columns is None:
        selected_fields = USER_FIELDS
    else:
        selected_fields = [f for f in USER_FIELDS if f.name in include_columns]

    headers = [f.label for f in selected_fields]
    ws.append(headers)

    for user in users:
        row = []

        for field in selected_fields:
            if field.name == "places_groups":
                value = get_places_groups_names(user)
            elif field.name == "user_groups":
                value = get_user_groups_names(user)
            else:
                value = get_field_value(user, field.getter)
                if value is None:
                    value = ""

            row.append(value)

        ws.append(row)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = None

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 15
    for col_num in range(2, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = 40

    return wb