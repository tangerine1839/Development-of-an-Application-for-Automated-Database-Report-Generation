from openpyxl import Workbook
from datetime import datetime
from typing import List, NamedTuple, Any, Optional

from openpyxl.styles import Alignment

from main.api_client.schemas.inspection_short import InspectionShort


class InspectionField(NamedTuple):
    name: str
    label: str
    getter: str


INSPECTION_FIELDS = [
    InspectionField("id", "ID инспекции", "id"),
    InspectionField("public_id", "Номер инспекции", "public_id"),
    InspectionField("status", "Статус", "status"),
    InspectionField("date", "Дата", "date"),
    InspectionField("place.id", "ID объекта", "place.id"),
    InspectionField("place.name", "Название объекта", "place.name"),
    InspectionField("place.address", "Адрес объекта", "place.address"),

    InspectionField("creator.id", "ID создателя", "creator.id"),
    InspectionField("creator.name", "Имя создателя", "creator.name"),
    InspectionField("creator.phone", "Телефон создателя", "creator.phone"),
    InspectionField("creator.email", "Email создателя", "creator.email"),

    InspectionField("assignee.id", "ID исполнителя", "assignee.id"),
    InspectionField("assignee.name", "Имя исполнителя", "assignee.name"),
    InspectionField("assignee.phone", "Телефон исполнителя", "assignee.phone"),
    InspectionField("assignee.email", "Email исполнителя", "assignee.email"),

    InspectionField("guest_link", "Ссылка", "guest_link"),
    InspectionField("is_guest_inspection", "Гостевая инспекция", "is_guest_inspection"),
]


def get_field_value(inspection: InspectionShort, field_path: str) -> Any:

    parts = field_path.split(".")
    current = inspection
    for part in parts:
        if not hasattr(current, part):
            return None
        current = getattr(current, part)
        if current is None:
            break
    return current


def create_inspections_excel(
    inspections: List[InspectionShort],
    columns: Optional[List[str]] = None,
    heading: str = "Отчёт по инспекциям",
) -> Workbook:

    if columns is None:
        columns = [f.name for f in INSPECTION_FIELDS]

    wb = Workbook()
    ws = wb.active
    ws.title = "Инспекции"

    headers = [f.label for f in INSPECTION_FIELDS if f.name in columns]
    ws.append(headers)

    field_paths = [f.getter for f in INSPECTION_FIELDS if f.name in columns]

    for inspection in inspections:
        row = []
        for path in field_paths:
            value = get_field_value(inspection, path)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif value is None:
                value = ""
            row.append(value)
        ws.append(row)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = None

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 50
    for col_num in range(2, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = 40

    return wb