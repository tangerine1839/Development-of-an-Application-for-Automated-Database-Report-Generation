from openpyxl import Workbook
from datetime import datetime, date
from typing import List, Dict, Any

from openpyxl.styles import Alignment, Font


class UserInspectionField:
    def __init__(self, name: str, label: str, getter: str):
        self.name = name
        self.label = label
        self.getter = getter


USER_INSPECTION_FIELDS = [
    UserInspectionField("user_id", "ID пользователя", "user_id"),
    UserInspectionField("user_name", "Имя пользователя", "user_name"),
    UserInspectionField("position", "Должность", "position"),
    UserInspectionField("inspections_count", "Количество проверок", "inspections_count"),
]


def get_field_value(item: Dict[str, Any], field_path: str) -> Any:
    parts = field_path.split(".")
    current = item
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        else:
            if not hasattr(current, part):
                return None
            current = getattr(current, part)
        if current is None:
            break
    return current


def create_users_inspections_count_excel(
        data: List[Dict[str, Any]],
        date_from: date,
        date_to: date,
        sorting: str,
        heading: str = "Отчёт по количеству проверок у пользователей",
) -> Workbook:


    wb = Workbook()
    ws = wb.active
    ws.title = "Проверки пользователей"

    ws.append([heading])
    ws.append([f"Период: с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"])
    sort_text = "по убыванию" if sorting == 'desc' else "по возрастанию"
    ws.append([f"Сортировка: {sort_text}"])
    ws.append([])

    headers = [f.label for f in USER_INSPECTION_FIELDS]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    total_inspections = 0
    for item in data:
        user_id = get_field_value(item, "user_id")
        user_name = get_field_value(item, "user_name")
        position = get_field_value(item, "position") or ""
        inspections_count = get_field_value(item, "inspections_count")

        total_inspections += inspections_count

        row = [user_id, user_name, position, inspections_count]
        ws.append(row)

        current_row = ws.max_row
        ws.row_dimensions[current_row].height = None

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            if col_num == 1 or col_num == 4:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws.append([])
    ws.append(["ИТОГО:", "", "", total_inspections])

    total_row = ws.max_row
    for col_num in range(1, 5):
        cell = ws.cell(row=total_row, column=col_num)
        cell.font = Font(bold=True)
        if col_num == 4:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

    return wb