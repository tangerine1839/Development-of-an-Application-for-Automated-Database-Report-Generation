from openpyxl import Workbook
from datetime import datetime
from typing import List, get_type_hints, NamedTuple, Any, Optional

from openpyxl.styles import Alignment

from main.api_client.schemas.task_short import TaskShort
from main.api_client.schemas.user_short import UserShort


class TaskField(NamedTuple):
    name: str
    label: str
    getter: str


TASK_FIELDS = [
    TaskField("id", "ID задачи", "id"),
    TaskField("public_id", "Номер задачи", "public_id"),
    TaskField("inspection_id", "Номер проверки", "inspection_id"),
    TaskField("status", "Статус", "status"),
    TaskField("title", "Название", "title"),
    TaskField("description", "Описание", "description"),
    TaskField("creator.id", "ID создателя", "creator.id"),
    TaskField("creator.name", "Имя создателя", "creator.name"),
    TaskField("creator.email", "Email создателя", "creator.email"),
    TaskField("creator.role", "Роль создателя", "creator.role"),
    TaskField("creator.position", "Должность создателя", "creator.position"),
    TaskField("creator.phone", "Телефон создателя", "creator.phone"),
    TaskField("assignee.id", "ID исполнителя", "assignee.id"),
    TaskField("assignee.name", "Имя исполнителя", "assignee.name"),
    TaskField("assignee.email", "Email исполнителя", "assignee.email"),
    TaskField("assignee.role", "Роль исполнителя", "assignee.role"),
    TaskField("assignee.position", "Должность исполнителя", "assignee.position"),
    TaskField("assignee.phone", "Телефон исполнителя", "assignee.phone"),
    TaskField("validator.id", "ID проверяющего", "validator.id"),
    TaskField("validator.name", "Имя проверяющего", "validator.name"),
    TaskField("validator.email", "Email проверяющего", "validator.email"),
    TaskField("validator.role", "Роль проверяющего", "validator.role"),
    TaskField("validator.position", "Должность проверяющего", "validator.position"),
    TaskField("validator.phone", "Телефон проверяющего", "validator.phone"),
    TaskField("priority", "Приоритет", "priority"),
    TaskField("created_at", "Дата создания", "created_at"),
    TaskField("expire_at", "Дата истечения", "expire_at"),
    TaskField("finished_at", "Дата завершения", "finished_at"),
]


def get_field_value(task: TaskShort, field_path: str) -> Any:

    parts = field_path.split(".")
    current = task
    for part in parts:
        if not hasattr(current, part):
            return None
        current = getattr(current, part)
        if current is None:
            break
    return current


def create_tasks_excel(
    tasks: List[TaskShort],
    columns: Optional[List[str]] = None,
    heading: str = "Отчёт по задачам",
) -> Workbook:

    if columns is None:
        columns = [f.name for f in TASK_FIELDS]

    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"

    headers = [f.label for f in TASK_FIELDS if f.name in columns]
    ws.append(headers)

    field_paths = [f.getter for f in TASK_FIELDS if f.name in columns]

    for task in tasks:
        row = []
        for path in field_paths:
            value = get_field_value(task, path)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif value is None:
                value = ""
            row.append(value)
        ws.append(row)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = None

        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 50
    for col_num in range(2, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = 40


    return wb